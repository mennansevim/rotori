"""Japonya gezi kanalı için içerik yol haritası + reels kataloğu Excel üretici.

Çıktı: content/Japonya_Reels_Roadmap.xlsx (6 sayfa).
İlke: her reel MERAK UYANDIRAN bir hook (soru/superlatif) ile başlar; detaylı
açıklama o sorunun BİLGİLENDİRİCİ cevabını verir. Uydurma sayı/fiyat yok.
"""
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# --- Marka renkleri ---
KIRMIZI = "8B1E1E"       # koyu kırmızı (kapak bandı)
KIRMIZI_ACIK = "C0392B"
LACIVERT = "1F2A44"
GRI = "F2F2F2"
GRI_KOYU = "D9D9D9"
BEYAZ = "FFFFFF"

thin = Side(style="thin", color="BBBBBB")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
WRAP = Alignment(wrap_text=True, vertical="top")
WRAP_CENTER = Alignment(wrap_text=True, vertical="center", horizontal="center")
HEAD_FONT = Font(bold=True, color=BEYAZ, size=11)
TITLE_FONT = Font(bold=True, size=16, color=KIRMIZI)
SUB_FONT = Font(italic=True, size=10, color="666666")


def style_header(ws, row, ncols, fill=KIRMIZI):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = PatternFill("solid", fgColor=fill)
        cell.font = HEAD_FONT
        cell.alignment = WRAP_CENTER
        cell.border = BORDER


def set_widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def write_rows(ws, start_row, rows, ncols, zebra=True):
    r = start_row
    for i, row in enumerate(rows):
        for c in range(1, ncols + 1):
            cell = ws.cell(row=r, column=c, value=row[c - 1] if c - 1 < len(row) else "")
            cell.alignment = WRAP
            cell.border = BORDER
            if zebra and i % 2 == 1:
                cell.fill = PatternFill("solid", fgColor=GRI)
        r += 1
    return r


wb = Workbook()

# =========================================================
# 00 BAŞLA
# =========================================================
ws = wb.active
ws.title = "00 Başla"
ws.sheet_properties.tabColor = KIRMIZI
set_widths(ws, [28, 90])
ws["A1"] = "🎬 Japonya Reels — İçerik Yol Haritası & Katalog"
ws["A1"].font = TITLE_FONT
ws.merge_cells("A1:B1")
ws["A2"] = "Sıfırdan 'en çok izlenen Japonya gezi rehberi' kanalı için üretim planı"
ws["A2"].font = SUB_FONT
ws.merge_cells("A2:B2")

basla = [
    ("Kanal", "@japonyaseyahat — Ailecek 13 gün Japonya (Tokyo, Osaka, Kyoto, Nara). Birinci ağız, samimi, 'biz gitmeden bilseydik' tonu."),
    ("Altın kural", "Her reel MERAK UYANDIRAN bir soru/iddia ile başlar (hook). Detaylı açıklama o sorunun BİLGİLENDİRİCİ cevabını verir. Uydurma sayı/fiyat yok."),
    ("Format ipucu", "İlk 1 saniyede güçlü hook (metin+görsel). Ekranda kısa not/altyazı. Trend ses. Sonda 'Kaydet/Takip et'. Seri kimliği kur."),
    ("Sayfalar", "01 Hedefler & KPI · 02 Kategoriler · 03 Reels Kataloğu (asıl) · 04 Paylaşım Takvimi · 05 Roadmap (faz faz)"),
    ("Durum kodları", "Fikir → Planlandı → Çekildi → Üretildi → Yayınlandı  (Katalogda açılır listeden seç)"),
    ("Öncelik", "1 = Elinde çekim var + yüksek etki (HEMEN) · 2 = Orta · 3 = Çekim/grafik gerekiyor (sonra)"),
    ("Çekim envanteri", "Tokyo Disneyland 73 · Nara 37 · Tokyo Havadan 37 · Kyoto 23 · Universal 21 · Osaka Havadan 12 · Fushimi Inari 6 · Osaka Kalesi 6 · Tokyo Tower 5 · Pokemon/Nintendo 3'er · Shinkansen 2 · tekil: Dotonbori, Konbini, Uniqlo, Shibuya, Japon Tuvaleti"),
    ("Nasıl kullanılır", "03 Reels Kataloğu'ndan Öncelik 1'leri seç → panelde (localhost:8420) ilgili kategoriden üret → Durum'u güncelle → 04 Takvim'e göre paylaş."),
]
r = 4
for k, v in basla:
    ws.cell(row=r, column=1, value=k).font = Font(bold=True)
    ws.cell(row=r, column=1).alignment = WRAP
    ws.cell(row=r, column=1).border = BORDER
    c = ws.cell(row=r, column=2, value=v)
    c.alignment = WRAP
    c.border = BORDER
    ws.row_dimensions[r].height = 42
    r += 1

# =========================================================
# 01 HEDEFLER & KPI
# =========================================================
ws = wb.create_sheet("01 Hedefler & KPI")
ws.sheet_properties.tabColor = LACIVERT
ws["A1"] = "Hedefler & KPI (Faz Faz)"
ws["A1"].font = TITLE_FONT
ws.merge_cells("A1:G1")
head = ["Faz", "Süre", "Takipçi Hedefi", "Ort. İzlenme Hedefi", "Save/Share Oranı", "Haftalık Reel", "Ana Strateji"]
for c, h in enumerate(head, 1):
    ws.cell(row=3, column=c, value=h)
style_header(ws, 3, len(head))
set_widths(ws, [16, 12, 16, 20, 16, 14, 46])
kpi = [
    ["Faz 0 – Kurulum", "1 hafta", "—", "—", "—", "3-9 (banka)", "Profil foto, bio, öne çıkanlar, ilk 9 grid, seri kimliği ve kapak şablonu hazırla."],
    ["Faz 1 – Isınma", "Ay 1-2", "0 → 1.000", "3.000-10.000", "%2+ save", "7", "Formatı bul: kısa, tek-bilgi, güçlü hook. Elindeki çekimlerle tutarlı üret; ne tutuyor gözle."],
    ["Faz 2 – İvme", "Ay 3-5", "1.000 → 10.000", "20.000-50.000", "%3+ save", "7", "Seri + trend ses + ekran altyazısı. 'Kaydet' değeri yüksek rehber içerik. Yorumlara reels ile cevap."],
    ["Faz 3 – Otorite", "Ay 6-12", "10.000 → 100.000", "100.000+", "%4+ save", "7-10", "İş birlikleri, tekrar-izlenmeye oynayan kurgu, öne çıkan serilerin devamı, mini rehber carousel'ler."],
]
r = write_rows(ws, 4, kpi, len(head))
ws.freeze_panes = "A4"
for rr in range(4, r):
    ws.row_dimensions[rr].height = 46

# =========================================================
# 02 KATEGORİLER
# =========================================================
ws = wb.create_sheet("02 Kategoriler")
ws.sheet_properties.tabColor = KIRMIZI_ACIK
ws["A1"] = "Konu Kategorileri"
ws["A1"].font = TITLE_FONT
ws.merge_cells("A1:G1")
head = ["Kategori", "Açıklama", "Amaç / Hedef Kitle", "Örnek Seri Adı", "Öncelik", "Çekim Durumu", "Not"]
for c, h in enumerate(head, 1):
    ws.cell(row=3, column=c, value=h)
style_header(ws, 3, len(head))
set_widths(ws, [26, 40, 28, 24, 9, 16, 30])
kategoriler = [
    ["Japonya'ya Gidiş Rehberi", "Gitmeden önce: ne zaman gidilir, vize, uçak, bütçe, JR Pass, internet, valiz.", "Planlama aşamasındaki yeni gezginler (en çok 'kaydet' alır)", "Gitmeden Bil", 1, "Çekilecek/Grafik", "Save canavarı; bio linkine rehber koy."],
    ["Varışta İlk Yapılacaklar", "İnince ilk 30 dk: IC kart, wifi, transfer, para.", "İlk kez gidenler", "İlk 24 Saat", 1, "Kısmi", "Havalimanı + konbini çekimi eklenebilir."],
    ["Shinkansen Rehberi", "Hızlı tren: koltuk, bilet, valiz, ekiben, Fuji.", "Şehirler arası gezenler", "Rayların Üstünde", 1, "Kısmi", "Elde Fuji + tren çekimi var."],
    ["Konbini & Yeme-İçme", "7-Eleven, onigiri, ATM, sokak lezzetleri.", "Herkes; pratik bilgi", "Konbini Sırları", 1, "Kısmi", "Tekil konbini çekimi var, çoğalt."],
    ["Kültür & Görgü Kuralları", "Tapınak adabı, tuvalet, bahşiş, sessizlik.", "Saygılı gezmek isteyenler", "Ayıp Olmasın", 2, "Kısmi", "Kısa 'yapma bunu' formatı iyi gider."],
    ["Alışveriş & Tax-Free", "Uniqlo, Yodobashi, Mitsui Outlet, Daiso, tax-free.", "Alışveriş sevenler", "Japonya'da Vur", 1, "Kısmi", "Uniqlo/Pokemon çekimi var."],
    ["Tokyo Rehberi", "Shibuya, Tokyo Tower, Skytree, teamLab, bedava manzara.", "Tokyo'ya gidenler", "Tokyo'da 6 Gün", 1, "Var", "Bol çekim mevcut."],
    ["Osaka Rehberi", "Dotonbori, Osaka Kalesi, yemek, havadan.", "Osaka'ya gidenler", "Osaka Lezzet", 1, "Var", "Kale + havadan + Dotonbori var."],
    ["Kyoto Rehberi", "Fushimi Inari, Gion, tapınaklar, ulaşım.", "Kültür/tarih sevenler", "Eski Japonya", 1, "Var", "23 klip + Fushimi var."],
    ["Nara & Geyikler", "Geyik kuralları, Todai-ji.", "Aileler, hayvan sevenler", "Nara Günü", 1, "Var", "37 klip, viral potansiyeli yüksek."],
    ["Tema Parkları", "Disney, Universal, Nintendo World taktikleri.", "Aileler, gençler", "Parkta Hayatta Kal", 1, "Var", "94+ klip; sıra/uygulama tüyoları."],
    ["Aile ile Japonya", "Çocuklarla rota, Shinkansen, çocuk dostu yerler.", "Ebeveynler", "Çocuklarla Japonya", 1, "Var", "Aile çekimleri güçlü."],
    ["Bütçe & Tasarruf", "Para nereye gider, ucuz numaralar.", "Bütçeli gezginler", "Cebini Koru", 2, "Grafik", "Genelde grafik/seslendirme."],
    ["Ulaşım & Uygulamalar", "Metro çıkışı, Google Maps, Suica, coin locker.", "Herkes", "Kaybolma", 2, "Kısmi", "Havadan + istasyon çekimi."],
    ["Sezon & Planlama", "12 aylık takvim, sakura, yağmur mevsimi.", "Ne zaman gideyim diyenler", "Hangi Ay?", 2, "Grafik", "Carousel/takvim formatı."],
    ["Rotalar & Süre", "Efsanevi 13 gün, günübirlik kaçamaklar, kaç gün.", "Rota kuranlar", "Efsane Rota", 1, "Var", "Havadan + şehir çekimleri."],
    ["Uçuş & Bilet", "Çin aktarmalı uçuşlar, ucuz bilet, aktarma.", "Bilet arayanlar", "Ucuza Uç", 2, "Grafik", "Güncel bilgi; kaynak göster."],
    ["Gizli Tüyolar", "Kayıp eşya, çöp, coin locker, sigara alanı, onsen.", "Deneyimli izleyici", "Kimse Söylemiyor", 2, "Karışık", "Merak-uyandıran kısa bilgiler."],
    ["Konaklama", "Business hotel, ryokan, oda neden küçük.", "Konaklama araştıranlar", "Nerede Kalmalı", 3, "Grafik", "Genelde stok/seslendirme."],
]
r = write_rows(ws, 4, kategoriler, len(head))
ws.freeze_panes = "A4"
dv_onc = DataValidation(type="list", formula1='"1,2,3"', allow_blank=True)
ws.add_data_validation(dv_onc)
dv_onc.add(f"E4:E{r-1}")
for rr in range(4, r):
    ws.row_dimensions[rr].height = 42

# =========================================================
# 03 REELS KATALOĞU
# =========================================================
ws = wb.create_sheet("03 Reels Kataloğu")
ws.sheet_properties.tabColor = "27AE60"
ws["A1"] = "Reels Kataloğu — Hook = Merak Sorusu · Açıklama = Bilgilendirici Cevap"
ws["A1"].font = TITLE_FONT
ws.merge_cells("A1:N1")
head = ["No", "Kategori", "Hook (Merak/Soru)", "Ekran Notu", "Açıklama (Bilgilendirici Cevap)",
        "Kullanılacak Görsel", "Format", "Ses/Müzik", "Hashtagler", "Öncelik", "Zorluk", "Durum", "Hafta", "Not"]
for c, h in enumerate(head, 1):
    ws.cell(row=3, column=c, value=h)
style_header(ws, 3, len(head))
set_widths(ws, [5, 20, 34, 18, 60, 20, 14, 14, 30, 8, 9, 12, 7, 20])

# (kategori, hook, ekran_notu, cevap, gorsel, format, ses, hashtags, oncelik, zorluk)
K = [
 # --- Gidiş Rehberi ---
 ("Gidiş Rehberi", "Japonya'ya en ucuz hangi ay gidilir?", "Takvimi kaydet 👇", "Zirve sezon sakura (mart sonu-nisan) ve sonbahar (kasım); fiyatlar bu dönemde tavan. En uygunu genelde ocak-şubat ile haziran (yağmur mevsimi). Hem hava hem fiyat dengesi istersen mayıs ve ekim (omuz sezonu) ideal.", "Stok/Grafik", "Tek Görsel+Not", "Sakin lofi", "japonya, japonyagezi, seyahat, traveltips, gezirehberi", 2, "Kolay"),
 ("Gidiş Rehberi", "Japonya'da internet: pocket wifi mi eSIM mi?", "Doğru seçim 👇", "Tek kişi ve kısa süre için eSIM en pratiği: uçakta inince anında aktif, ucuz. Aile veya çok cihaz varsa pocket wifi mantıklı: tek cihazdan herkes bağlanır. İkisini de havalimanından teslim-iade edebilirsin.", "Çekilecek", "Seslendirme", "Trend ses", "japonya, esim, pocketwifi, japonyagezi, traveltips", 2, "Orta"),
 ("Gidiş Rehberi", "JR Pass hâlâ mantıklı mı, yoksa para mı yakıyor?", "Hesabı yap 👇", "Zamlardan sonra JR Pass yalnızca çok sayıda şehir arası (ör. Tokyo-Osaka-Hiroshima) gezersen amorti ediyor. Sadece Tokyo-Osaka gidiş-dönüş yapacaksan tek tek bilet çoğu zaman daha ucuz. Rotanı yaz, tek tek fiyatla, sonra karar ver.", "Shinkansen & Fuji", "Seslendirme", "Voiceover (TR)", "jrpass, shinkansen, japonya, traveltips, japonyagezi", 2, "Orta"),
 ("Gidiş Rehberi", "Valizini nasıl hafiflettik? (kimse bilmiyor)", "Sırrı takkyubin 👇", "Şehirler arası geçişte valizi sırtlamadık: 'takkyubin' kargo ile valizi bir otelden diğerine önceden yolladık, trene elimiz boş bindik. Ertesi gün valiz otelde bizi bekliyordu.", "Shinkansen & Fuji", "Montaj", "Trend ses", "takkyubin, japonya, seyahattuyolari, japonyagezi, traveltips", 1, "Kolay"),
 ("Gidiş Rehberi", "Japonya'da en çok para nereye gidiyor?", "Bütçeyi buna göre kur 👇", "En büyük kalem uçak ve konaklama. Yeme-içme konbini ve lokal esnafla şaşırtıcı derecede makul. Asıl 'sürpriz' masraf ise şehirler arası Shinkansen biletleri ve tema parkı giriş+ekstraları. Bunları önceden bütçele.", "Stok/Grafik", "Tek Görsel+Not", "Sakin lofi", "japonyabutce, japonya, seyahat, traveltips, japonyagezi", 2, "Kolay"),
 # --- Varışta İlk Yapılacaklar ---
 ("Varışta İlk Yapılacaklar", "Japonya'ya inince ilk 30 dakikada ne yapmalısın?", "Sırayla 👇", "1) IC kart al (Suica/ICOCA/Welcome Suica). 2) eSIM/pocket wifi'yi aktive et. 3) İstersen valizi takkyubin ile otele yolla. 4) Havalimanı-şehir için doğru treni ya da limuzin otobüsü seç. Bu dördü ilk yarım saatini kurtarır.", "Konbini (7-Eleven)", "Seslendirme", "Voiceover (TR)", "japonya, ilkgun, tokyo, traveltips, japonyagezi", 1, "Orta"),
 ("Varışta İlk Yapılacaklar", "Suica nedir, neden ilk iş bu?", "Bir kart her yerde 👇", "Suica temassız bir IC kart: metro, otobüs, konbini, otomat — hepsinde bas geç. Bir kez yükle, bilet kuyruğuyla hiç uğraşma. iPhone'da Apple Wallet'a da eklenip telefondan kullanılabiliyor.", "Çekilecek", "Seslendirme", "Trend ses", "suica, japonya, ulasim, traveltips, japonyagezi", 2, "Orta"),
 ("Varışta İlk Yapılacaklar", "Narita mı Haneda mı? Şehre en kolay geçiş", "Havalimanına göre 👇", "Haneda şehre yakın (monoray/Keikyu ile ~30 dk). Narita uzak (Skyliner/N'EX ile ~1 saat). Gece geç indiysen ya da valizin çoksa limuzin otobüs en konforlusu. İnişten önce hangi hattı kullanacağını belirle.", "Tokyo Havadan", "Seslendirme", "Voiceover (TR)", "haneda, narita, tokyo, japonya, traveltips", 2, "Orta"),
 ("Varışta İlk Yapılacaklar", "Japonya'da kart mı geçer nakit mi?", "Cebinde ikisi de olsun 👇", "Büyük yerler kart/temassız kabul ediyor ama küçük lokanta, tapınak ve bazı otomatlar hâlâ nakit istiyor. Yabancı kartı en sorunsuz kabul eden nakit noktası 7-Eleven ATM'leri.", "Konbini (7-Eleven)", "Montaj", "Trend ses", "japonya, nakit, konbini, traveltips, japonyagezi", 1, "Kolay"),
 # --- Shinkansen ---
 ("Shinkansen Rehberi", "Shinkansen'de Fuji hangi koltuktan görünür?", "Sağ koltuk sırrı 👇", "Tokyo'dan Osaka yönüne giderken SAĞ taraftaki (D-E) koltukları seç; Fuji Dağı yolun o tarafından görünüyor. Dönüşte tam tersi, sol taraf. Rezervasyonu buna göre yaptır, cam kenarını kap.", "Shinkansen & Fuji", "Montaj", "Trend ses", "shinkansen, fuji, japonya, traveltips, bullettrain", 1, "Kolay"),
 ("Shinkansen Rehberi", "NOZOMI, HIKARI, KODAMA farkı ne?", "Doğru treni seç 👇", "NOZOMI en hızlısı, az durur (ama JR Pass'in eski sürümünde geçmez). HIKARI orta hız, birkaç durak. KODAMA her istasyonda durur, en yavaşı. Acelen yoksa HIKARI hız-fiyat dengesi için ideal.", "Shinkansen & Fuji", "Seslendirme", "Voiceover (TR)", "shinkansen, nozomi, japonya, traveltips, japonyagezi", 2, "Orta"),
 ("Shinkansen Rehberi", "Shinkansen'de büyük valiz nereye konur?", "Önceden rezerve et 👇", "Küçük valiz üst rafa sığar. Büyük valizin varsa vagonun en arka sırasındaki 'oversized baggage' koltuğunu önceden rezerve etmen gerekiyor; yoksa yer bulamayabilirsin.", "Shinkansen & Fuji", "Seslendirme", "Voiceover (TR)", "shinkansen, valiz, japonya, traveltips, japonyagezi", 3, "Orta"),
 ("Shinkansen Rehberi", "Trende 'ekiben' nedir, neden denemelisin?", "Rayların lezzeti 👇", "Ekiben, istasyonlarda satılan bölgeye özel kutu yemek. Trene binmeden al, cam kenarında manzara eşliğinde ye. Her bölgenin kendi ekiben'i var; deneyimin bir parçası.", "Çekilecek", "Montaj", "Sakin lofi", "ekiben, shinkansen, japonyayemek, traveltips, japonyagezi", 3, "Kolay"),
 # --- Nara ---
 ("Nara & Geyikler", "Nara geyikleri neden başını eğiyor?", "Cevap şaşırtıcı 👇", "Doğuştan değil, öğrenilmiş bir davranış. Nara'nın geyikleri kutsal sayıldığı için nesillerdir korunuyor ve ziyaretçilerden bisküvi (shika senbei) almayı öğrenmişler. Sen selam verince onlar da baş eğip bisküviyi bekliyor; bu, insanlarla kurdukları 'alışverişin' sonucu.", "Nara Geyikleri", "Montaj", "Trend ses", "nara, geyik, japonya, narageyikleri, traveltips", 1, "Kolay"),
 ("Nara & Geyikler", "Nara geyiklerine ne yedirebilirsin?", "Sadece bunu ver 👇", "Yalnızca parktaki resmi tezgahlarda satılan 'shika senbei' bisküvisini. Kendi getirdiğin atıştırmalığı verme; onlara zararlı ve tepki çekiyor. Bisküvi bittiğinde avucunu açıp göster, 'bitti' anlıyorlar.", "Nara Geyikleri", "Montaj", "Trend ses", "nara, shikasenbei, geyik, japonya, traveltips", 1, "Kolay"),
 ("Nara & Geyikler", "Nara'da geyik ısırır mı? Bu hatayı yapma", "Cebine dikkat 👇", "Bisküviyi cebinde ya da poşette saklarsan kokusunu alıp üstüne gelir, kıyafetini çekiştirebilir. Bisküviyi elinde tut, verirken avucunu düz aç. Küçük çocukları yanında tut.", "Nara Geyikleri", "Montaj", "Enerjik trend", "nara, geyik, japonya, aileilegezi, traveltips", 1, "Kolay"),
 # --- Kyoto ---
 ("Kyoto Rehberi", "Kyoto'da metro neden işe yaramaz?", "Otobüsü öğren 👇", "Kyoto metrosu çok sınırlı; şehir otobüs ağırlıklı ve tapınakların çoğu otobüs hattında. Metroya güvenip rota kurarsan vakit kaybedersin. Otobüs + kısa yürüyüş kombinini planla.", "Kyoto", "Seslendirme", "Voiceover (TR)", "kyoto, ulasim, japonya, kyototravel, traveltips", 1, "Kolay"),
 ("Kyoto Rehberi", "Fushimi Inari'nin torii'lerinde ne yazıyor?", "Kimse fark etmiyor 👇", "O binlerce kırmızı torii aslında bağış yapan kişi ve şirketlerin adına dikilmiş. Arka yüzlerine bağışçının adı ve tarih yazılıyor. Yani her kapı bir bağış; bir dahaki sefere arkalarına bak.", "Kyoto Fushimi Inari", "Montaj", "Sakin lofi", "fushimiinari, kyoto, torii, japonya, traveltips", 1, "Kolay"),
 ("Kyoto Rehberi", "Fushimi Inari'de kalabalıktan nasıl kaçarsın?", "Yukarı çık 👇", "Kalabalık hep girişte toplanıyor. Patikadan yukarı tırmandıkça insanlar eriyor, tepede neredeyse yalnız kalıyorsun. Giriş ücretsiz ve 24 saat açık; erken ya da geç saat en tenhası.", "Kyoto Fushimi Inari", "Montaj", "Sakin lofi", "fushimiinari, kyoto, japonya, kyototravel, traveltips", 1, "Kolay"),
 ("Kyoto Rehberi", "Gion'da geyşa fotoğrafı çekmek yasak mı?", "Ceza yiyebilirsin 👇", "Gion'un bazı özel ara sokaklarında geyşaların izinsiz fotoğrafını çekmek yasak ve para cezası var; sokak başlarında uyarı tabelaları asılı. Ana caddede sorun yok, ama özel sokaklarda kurallara uy.", "Kyoto", "Seslendirme", "Voiceover (TR)", "gion, kyoto, japonya, geysha, traveltips", 2, "Orta"),
 ("Kyoto Rehberi", "Tapınakta el yıkama nasıl yapılır? (temizu)", "Sırası önemli 👇", "Girişteki çeşmede maşrapayla: önce sol elini, sonra sağ elini yıka, ardından avucuna su alıp ağzını çalkala (maşrapaya değdirmeden). En son maşrapayı dik tutup sapını akıt. Böylece arınıp giriyorsun.", "Kyoto Tapınakları", "Seslendirme", "Voiceover (TR)", "kyoto, tapinak, temizu, japonya, gorgukurallari", 2, "Orta"),
 # --- Osaka ---
 ("Osaka Rehberi", "Dotonbori'de o meşhur kareyi nerede çekersin?", "Tam bu köprüde 👇", "Meşhur Glico koşan adam tabelasının en temiz karesi Ebisubashi köprüsünün üstünden çıkıyor. Işıklar asıl gece açılıyor; gündüz aynı yer sönük kalıyor, o yüzden akşam git.", "Dotonbori", "Montaj", "Enerjik trend", "dotonbori, osaka, glico, japonya, traveltips", 1, "Kolay"),
 ("Osaka Rehberi", "Osaka Kalesi'nin içinde ne var?", "Sadece dışı değil 👇", "Kalenin içi tamamen müzeye çevrilmiş ve en üst kata çıkınca şehir manzarası açılıyor. Kaleyi çevreleyen park çok geniş; ana kapıdan kaleye yürüyüş hatırı sayılır sürüyor, zamanını ona göre ayır.", "Osaka Kalesi", "Montaj", "Sakin lofi", "osakacastle, osaka, japonya, osakakalesi, traveltips", 1, "Kolay"),
 ("Osaka Rehberi", "Osaka mı Tokyo mu daha lezzetli?", "Cevap net 👇", "Osaka'ya 'Japonya'nın mutfağı' deniyor. Takoyaki, okonomiyaki, kushikatsu gibi sokak lezzetlerinin başkenti burası. Tokyo daha rafine restoranlarıyla öne çıkıyor; sokak yemeği için Osaka.", "Dotonbori", "Montaj", "Enerjik trend", "osaka, takoyaki, japonyayemek, japonya, traveltips", 2, "Kolay"),
 ("Osaka Rehberi", "Osaka'yı gökten görünce ne değişiyor?", "Şehri baştan tanı 👇", "Havadan bakınca Osaka'nın kanalları ve kalenin hendekleri şehrin dokusunu bambaşka anlatıyor. Ama dikkat: Japonya'da drone kuralları çok sıkı, kalabalık ve tapınak üstünde uçmak yasak; izinli açık alan seç.", "Osaka Havadan", "Montaj", "Sakin lofi", "osaka, drone, havadan, japonya, traveltips", 2, "Orta"),
 # --- Tokyo ---
 ("Tokyo Rehberi", "Tokyo'da en iyi manzara bedava — nerede?", "Kimse bilmiyor 👇", "Tokyo Metropolitan Government binasının gözlem katı ücretsiz ve şehri kuşbakışı görüyorsun. Shibuya Sky ve Skytree ücretli; ama bedava manzara isteyen buraya gitsin.", "Tokyo Havadan", "Seslendirme", "Voiceover (TR)", "tokyo, manzara, bedava, japonya, traveltips", 1, "Orta"),
 ("Tokyo Rehberi", "Shibuya Crossing'i yukarıdan nasıl çekersin?", "İki nokta var 👇", "En iyi kuşbakışı kare için iki yer: Shibuya Sky teras katı ya da istasyondaki Starbucks'ın üst katı. Aşağıdan çekeceksen ışık döngüsünde herkesin aynı anda geçtiği anı yakala.", "Shibuya Meydanı", "Montaj", "Trend ses", "shibuya, tokyo, shibuyacrossing, japonya, traveltips", 1, "Kolay"),
 ("Tokyo Rehberi", "Tokyo Tower mı Skytree mi?", "Karar rehberi 👇", "Skytree daha yüksek ve modern, manzara daha uzağa uzanıyor. Tokyo Tower nostaljik ve turuncu-beyaz yapısı özellikle dipten fotoğrafta efsane duruyor. Manzara için Skytree, estetik kare için Tokyo Tower.", "Tokyo Tower", "Montaj", "Sakin lofi", "tokyotower, skytree, tokyo, japonya, traveltips", 2, "Kolay"),
 ("Tokyo Rehberi", "Tokyo metrosunda yanlış çıkış seni nasıl yakar?", "Numarayı takip et 👇", "Shinjuku, Shibuya gibi dev istasyonlarda yanlış çıkış 10-15 dakika kaybettirir. Google Maps sana çıkış numarasını (ör. A5) veriyor; tabeladaki harf-numarayı takip edip doğru çıkıştan çık.", "Tokyo Havadan", "Seslendirme", "Voiceover (TR)", "tokyo, metro, ulasim, japonya, traveltips", 2, "Orta"),
 ("Tokyo Rehberi", "teamLab neden bu kadar konuşuluyor?", "Başka dünya 👇", "teamLab, gezilebilen dijital bir sanat müzesi: su dolu odalar, ışık tünelleri, aynalı sonsuzluklar. Çocuk da yetişkin de büyüleniyor. Bileti önceden almazsan giremeyebilirsin.", "Çekilecek", "Montaj", "Trend ses", "teamlab, tokyo, japonya, sanat, traveltips", 3, "Orta"),
 # --- Tema Parkları ---
 ("Tema Parkları", "Tokyo Disney'de sırada beklememenin yolu?", "Uygulama şart 👇", "Park girişinde resmi uygulamadan 'Premier Access' alınca popüler binitlerde sıra beklemiyorsun. Bazı binitlere ise Standby veya Priority Pass'i yine uygulamadan almadan giremiyorsun. İlk iş uygulamayı indir.", "Tokyo Disneyland", "Montaj", "Enerjik trend", "tokyodisney, disneyland, tokyo, japonya, traveltips", 1, "Orta"),
 ("Tema Parkları", "Universal Nintendo World'e nasıl girilir?", "Timed entry sırrı 👇", "Nintendo World'e giriş çoğu zaman zaman-bantlı. Uygulamadan 'Area Timed Entry' almadan alana giremeyebiliyorsun. Power-Up Band alırsan bloklara vurup etkileşimli mini oyunlar oynuyorsun.", "Universal-Nintendo World", "Montaj", "Enerjik trend", "nintendoworld, usj, osaka, supermario, traveltips", 1, "Orta"),
 ("Tema Parkları", "Universal'de Express Pass şart mı?", "Duruma bağlı 👇", "Express Pass biniş sırasını ciddi kısaltıyor ama ayrı ve pahalı. Kalabalık günlerde (tatil/hafta sonu) işe yarıyor; sakin bir gün gidiyorsan gerek yok. Önce takvime bak, sonra al.", "Universal Studios Japan", "Seslendirme", "Voiceover (TR)", "universalstudios, usj, osaka, japonya, traveltips", 2, "Orta"),
 ("Tema Parkları", "Disney'de popcorn kovası neden bu kadar önemli?", "Koleksiyon işi 👇", "Popcorn kovalarının tasarımları sınırlı ve koleksiyonluk; sevdiğini gördüğün an al, sonra o tasarımı bulamıyorsun. Aynı zamanda park içinde pratik atıştırmalık taşıma yöntemi.", "Tokyo Disneyland", "Montaj", "Trend ses", "tokyodisney, popcorn, disneyland, japonya, traveltips", 3, "Kolay"),
 # --- Konbini & Yeme-İçme ---
 ("Konbini & Yeme-İçme", "Konbini'de bütün geziyi kurtaran 3 şey?", "Not al 👇", "1) 7-Bank ATM: yabancı kartla nakit çekmenin en sorunsuz yolu. 2) Onigiri: ucuz, doyurucu, taze. 3) Sıcak yemekleri kasada 'atatamemasu ka?' deyince ısıtıyorlar. Konbini sadece market değil, kurtarıcı.", "Konbini (7-Eleven)", "Montaj", "Trend ses", "konbini, 7eleven, japonya, traveltips, japonyagezi", 1, "Kolay"),
 ("Konbini & Yeme-İçme", "Onigiri paketi nasıl açılır?", "1-2-3 sırası 👇", "Paketin üstünde numaralı 3 adım var. Sırayla çekersen yosun (nori) pirinçten ayrı, gevrek kalıyor ve tam sarınca ıslanmadan yiyorsun. Yanlış açarsan yosun ezilip dağılıyor.", "Konbini (7-Eleven)", "Montaj", "Trend ses", "onigiri, konbini, japonyayemek, japonya, traveltips", 1, "Kolay"),
 ("Konbini & Yeme-İçme", "Japonya'da su almana gerek var mı?", "Cebinde kalsın 👇", "Musluk suyu içilebilir ve her yerde temiz. Otelde termosunu doldur, gün boyu şişe suya para verme. Sıcakta bile bu küçük alışkanlık ciddi tasarruf.", "Çekilecek", "Tek Görsel+Not", "Sakin lofi", "japonya, tasarruf, traveltips, japonyagezi, seyahat", 2, "Kolay"),
 ("Konbini & Yeme-İçme", "Japonya'da bahşiş veriliyor mu?", "Sakın verme 👇", "Hayır. Japonya'da bahşiş kültürü yok; bırakmaya çalışmak kafa karıştırıcı, hatta kırıcı olabiliyor. Servis zaten fiyata dahil ve kalite yüksek. Parayı masada bırakma.", "Stok/Grafik", "Tek Görsel+Not", "Voiceover (TR)", "japonya, bahsis, gorgukurallari, traveltips, japonyagezi", 2, "Kolay"),
 ("Konbini & Yeme-İçme", "Ramen otomatından nasıl sipariş verilir?", "Korkma, kolay 👇", "Birçok ramen dükkanında girişte bir fiş makinesi var: paranı at, istediğin ramenin tuşuna bas, çıkan fişi tezgaha ver. İngilizce konuşmana, para üstü derdine gerek kalmıyor.", "Çekilecek", "Seslendirme", "Trend ses", "ramen, japonyayemek, japonya, traveltips, japonyagezi", 2, "Kolay"),
 # --- Kültür & Görgü ---
 ("Kültür & Görgü Kuralları", "Japon tuvaletindeki büyük düğme ne işe yarar?", "Yanlışına basma 👇", "O büyük düğme genelde sifon değil, bide. Panelde ayrıca 'ses' düğmesi olur: su sesi çıkarıp mahremiyet sağlıyor. Sifon çoğu zaman duvarda ya da kolda. Panele bir saniye bak, sonra bas.", "Japon Tuvaletleri", "Montaj", "Enerjik trend", "japonya, japontuvaleti, kultur, traveltips, japonyagezi", 1, "Kolay"),
 ("Kültür & Görgü Kuralları", "Japonya'da yürürken yemek neden ayıp?", "Buna dikkat 👇", "Japonya'da genelde yürürken yenmez; aldığın yerde ya da bir kenarda durup yenir. Ayrıca sokakta çöp kutusu çok az, çöpünü otele/konbini'ye kadar taşıman gerekebilir.", "Stok/Grafik", "Seslendirme", "Voiceover (TR)", "japonya, gorgukurallari, kultur, traveltips, japonyagezi", 2, "Kolay"),
 ("Kültür & Görgü Kuralları", "Tapınakta dua nasıl edilir?", "Sıra şöyle 👇", "Sunağın önünde: hafif reverans, parayı at, (varsa) çanı çal, sonra 2 kez eğil, 2 kez el çırp, dilek tut, 1 kez daha eğil. İçeride sessizlik ve fotoğraf yasağına dikkat et.", "Kyoto Tapınakları", "Seslendirme", "Voiceover (TR)", "japonya, tapinak, gorgukurallari, kultur, traveltips", 2, "Orta"),
 ("Kültür & Görgü Kuralları", "Toplu taşımada neden herkes sessiz?", "Telefonu kaldır 👇", "Japonya'da metro/tren içinde telefonla konuşulmaz, sesli video izlenmez, telefon sessizdedir. Sessizlik saygı göstergesi. Konuşman gerekiyorsa alçak sesle ve kısa.", "Tokyo Havadan", "Tek Görsel+Not", "Sakin lofi", "japonya, kultur, gorgukurallari, traveltips, japonyagezi", 3, "Kolay"),
 # --- Alışveriş ---
 ("Alışveriş & Tax-Free", "Japonya'nın en büyük outlet'i hangisi?", "Cevap: Mitsui Outlet 👇", "Mitsui Outlet Park zinciri Japonya'nın en yaygın ve büyük outlet'lerinden; yüzlerce yerli-yabancı marka ve tax-free imkanı var. Şehir dışında ama trenle/otobüsle ulaşımı kolay. Alışveriş odaklıysan bir gün ayır.", "Çekilecek/Stok", "Seslendirme", "Trend ses", "mitsuioutlet, japonya, alisveris, outlet, traveltips", 2, "Orta"),
 ("Alışveriş & Tax-Free", "Japonya'da Uniqlo'yu neden orada vurmalısın?", "3 sebep 👇", "1) Büyük şubelerde pasaportla anında tax-free. 2) Japonya'ya özel UT tişört tasarımları sadece orada. 3) Aldığın pantolonun paçasını ücretsiz ve dakikalar içinde kısaltıyorlar. Bavul yerin varsa kaçırma.", "Uniqlo", "Montaj", "Trend ses", "uniqlo, japonya, taxfree, alisveris, traveltips", 1, "Kolay"),
 ("Alışveriş & Tax-Free", "Yodobashi ve Bic Camera nedir?", "Teknoloji cenneti 👇", "Kat kat kamera, lens, gadget, oyuncak satan dev elektronik mağazaları. Tax-free var, fiyatlar rekabetçi. Teknoloji seviyorsan saatlerini burada geçirebilirsin; biz de öyle yaptık.", "Çekilecek", "Montaj", "Enerjik trend", "yodobashi, biccamera, japonya, teknoloji, traveltips", 2, "Orta"),
 ("Alışveriş & Tax-Free", "Pokemon Center'da neyi kaçırma?", "Şubeye özel 👇", "Her Pokemon Center'da sadece o şubeye/şehre özel ürünler satılıyor, hepsi her yerde yok. Büyük şubelerde tax-free var ve kasada hediye paketi ücretsiz, çok özenli.", "Pokemon Center", "Montaj", "Trend ses", "pokemon, pokemoncenter, japonya, tokyo, traveltips", 1, "Kolay"),
 ("Alışveriş & Tax-Free", "100 yen mağazaları (Daiso) neden hazine?", "Hediyelik burada 👇", "Daiso gibi 100 yen mağazalarında kaliteli hediyelik ve pratik ürünler çok ucuz. Dönüşte herkese hediye almak için birebir; hem ucuz hem 'çok Japonya'.", "Çekilecek", "Montaj", "Trend ses", "daiso, 100yen, japonya, alisveris, traveltips", 3, "Kolay"),
 ("Alışveriş & Tax-Free", "Tax-free nasıl yapılır, nelere dikkat?", "Adım adım 👇", "Kasada pasaportunu göster, aynı mağazada eşik tutarı geç, çoğu yerde ayrı 'tax-free' kasası var. Vergisiz aldığın ürünü Japonya'da açman yasak, mühürlü paketle çıkış yaparsın.", "Uniqlo", "Seslendirme", "Voiceover (TR)", "taxfree, japonya, alisveris, traveltips, japonyagezi", 2, "Orta"),
 # --- Aile ---
 ("Aile ile Japonya", "Çocuklarla Tokyo'da ne yapılır?", "Unutulmaz liste 👇", "teamLab (dijital sanat), Ueno Doğa Bilimleri Müzesi (dinozor iskeletleri), Odaiba (Gundam), ve tabii Disneyland. Yürüme mesafeleri uzun, aralara mola ve konbini atıştırması koy.", "Tokyo Disneyland", "Montaj", "Enerjik trend", "cocuklarlatokyo, aileilegezi, tokyo, japonya, traveltips", 1, "Orta"),
 ("Aile ile Japonya", "Çocukla Shinkansen zor mu?", "Aksine çok kolay 👇", "Shinkansen çocukla seyahat için ideal: geniş koltuk, temiz tuvalet, sessiz ortam. Büyük valiz için oversized koltuğu önceden rezerve et; çocuk cam kenarında Fuji'yi izlesin.", "Shinkansen & Fuji", "Montaj", "Sakin lofi", "aileilegezi, shinkansen, cocuklajaponya, japonya, traveltips", 2, "Kolay"),
 ("Aile ile Japonya", "Japonya çocuk dostu bir ülke mi?", "Cevap: çok 👇", "Evet. Tertemiz tuvaletler, bebek bakım odaları, her köşede konbini, düşük suç oranı. Çocukla dolaşmak şaşırtıcı derecede rahat; toplu taşıma bile stressiz.", "Nara Geyikleri", "Tek Görsel+Not", "Sakin lofi", "aileilegezi, japonya, cocuklajaponya, traveltips, japonyagezi", 3, "Kolay"),
 # --- Ulaşım & Uygulamalar ---
 ("Ulaşım & Uygulamalar", "Japonya'da hangi uygulamalar şart?", "İndir gel 👇", "Google Maps (toplu taşıma tarifleri kusursuz), Apple Wallet'ta Suica, çeviri için Google Translate ya da Papago, ve tema parkları için resmi uygulamalar. Bu dördü olmadan yola çıkma.", "Stok/Grafik", "Seslendirme", "Voiceover (TR)", "japonya, uygulama, ulasim, traveltips, japonyagezi", 2, "Kolay"),
 ("Ulaşım & Uygulamalar", "Google Maps Japonya'da neden bu kadar iyi?", "Perona kadar söyler 👇", "Japonya'da Google Maps sana sadece hangi tren değil; hangi peron, hangi vagon, hangi çıkış numarası ve aktarma dakikasını da veriyor. Harfiyen uyarsan hiç kaybolmuyorsun.", "Tokyo Havadan", "Seslendirme", "Voiceover (TR)", "googlemaps, japonya, ulasim, traveltips, japonyagezi", 3, "Kolay"),
 ("Ulaşım & Uygulamalar", "Valizle gezme: coin locker sırrı", "Elin boş gezsin 👇", "İstasyonlardaki kilitli dolaplara (coin locker) valizini bırak, şehri elin boş gez. Çoğu IC kartla açılıp kapanıyor; otel check-in'ini beklerken bavulla dolaşmıyorsun.", "Çekilecek", "Montaj", "Trend ses", "coinlocker, japonya, ulasim, traveltips, japonyagezi", 2, "Kolay"),
 # --- Sezon & Planlama ---
 ("Sezon & Planlama", "Japonya'da 12 ay: hangi ay ne var?", "Takvimi kaydet 👇", "Kısa özet: mart sonu-nisan sakura, temmuz-ağustos festival ve nem, kasım kızıl yapraklar (koyo), aralık kış ışıklandırmaları. Gideceğin aya göre hem manzara hem kıyafet değişiyor.", "Stok/Grafik", "Carousel", "Sakin lofi", "japonya, seyahattakvimi, sakura, japonyagezi, traveltips", 2, "Orta"),
 ("Sezon & Planlama", "Kiraz çiçeği (sakura) ne zaman açar?", "Kaçırma 👇", "Bölgeye göre değişse de genelde mart sonu-nisan başı. Açılış tahmini (sakura zensen) yayınlanıyor ve tam çiçeklenme sadece birkaç gün sürüyor; tarihini buna göre esnek tut.", "Kyoto", "Tek Görsel+Not", "Sakin lofi", "sakura, japonya, kirazicicegi, japonyagezi, traveltips", 2, "Kolay"),
 ("Sezon & Planlama", "Japonya'da yağmur mevsimi (tsuyu) ne zaman?", "Artı ve eksileri 👇", "Genelde haziran-temmuz başı nemli ve yağışlı geçer. Kötü yanı yağmur; iyi yanı fiyatlar düşer, yerler tenhalaşır. Gidersen şemsiye/yağmurluk şart, planı esnek tut.", "Stok/Grafik", "Seslendirme", "Voiceover (TR)", "japonya, yagmurmevsimi, tsuyu, japonyagezi, traveltips", 3, "Kolay"),
 # --- Rotalar & Süre ---
 ("Rotalar & Süre", "13 günde Japonya: efsane rota nasıl kurulur?", "Rotayı kaydet 👇", "Bizim rota: Tokyo'da 6 gece (Ikebukuro üs) → Shinkansen ile Osaka'ya geçiş → Osaka üs, buradan Kyoto ve Nara günübirlik → Osaka'dan dönüş. İlk kez gidenler için dengeli ve yorucu olmayan bir akış.", "Tokyo Havadan", "Carousel", "Trend ses", "japonyarota, japonya, 13gun, japonyagezi, traveltips", 1, "Orta"),
 ("Rotalar & Süre", "Osaka'dan günübirlik gidilecek 6 rota?", "Hepsi trenle yakın 👇", "Osaka'yı üs yaparsan Kyoto, Nara, Kobe, Himeji (kale), Uji (matcha) ve Wakayama'ya trenle 30-90 dakikada gidip dönebilirsin. Otel değiştirmeden 6 farklı şehir.", "Osaka Kalesi", "Carousel", "Trend ses", "osaka, gunubirlik, japonyarota, japonya, traveltips", 2, "Orta"),
 ("Rotalar & Süre", "Japonya'ya kaç gün yeter?", "Cevap 👇", "İlk kez için 10-14 gün ideal: Tokyo + Osaka + Kyoto + Nara rahat sığar, yorulmadan gezirsin. Daha kısa sürede tek bölgeye (ör. sadece Tokyo çevresi) odaklan.", "Tokyo Havadan", "Tek Görsel+Not", "Voiceover (TR)", "japonya, kacgun, japonyarota, japonyagezi, traveltips", 2, "Kolay"),
 # --- Uçuş & Bilet ---
 ("Uçuş & Bilet", "Çin aktarmalı Japonya uçuşları: nelere dikkat?", "Ucuz ama... 👇", "Çin aktarmalı biletler genelde en ucuzu ama aktarma süresi, transit vize/TWOV kuralları ve bagajın direkt gidip gitmediği değişebiliyor. Uzun aktarmalarda kuralları havayolundan güncel olarak teyit et; ucuzluk uğruna sürprize hazır ol.", "Çekilecek/Stok", "Seslendirme", "Voiceover (TR)", "japonya, ucakbileti, aktarma, seyahat, traveltips", 2, "Orta"),
 ("Uçuş & Bilet", "Japonya'ya en ucuz bileti nasıl bulursun?", "3 taktik 👇", "1) Tarihte esnek ol (hafta içi ve sezon dışı ucuz). 2) Aktarmalı uçuşlara bak. 3) Fiyat takip alarmı kur, düşünce yakala. Erken planla ama son dakika fırsatlarını da izle.", "Stok/Grafik", "Tek Görsel+Not", "Trend ses", "ucakbileti, japonya, ucuzucus, seyahat, traveltips", 3, "Kolay"),
 # --- Gizli Tüyolar ---
 ("Gizli Tüyolar", "Japonya'da kaybettiğin cüzdan geri gelir mi?", "İnanılmaz ama 👇", "Genelde evet. Japonya'da kayıp eşya kültürü çok güçlü; bulunan cüzdan/telefon en yakın koban'a (küçük polis kulübesi) ya da istasyon görevlisine teslim ediliyor. Kaybedersen önce oraya sor.", "Stok/Grafik", "Seslendirme", "Trend ses", "japonya, gizlituyolar, kayipeşya, japonya, traveltips", 3, "Kolay"),
 ("Gizli Tüyolar", "Japonya'da neden hiç çöp kutusu yok?", "Sebebi ilginç 👇", "Sokakta bilinçli olarak çok az çöp kutusu var; halk çöpünü eve/otele kadar taşıyor ve titizce ayrıştırıyor. Yanında küçük bir poşet taşı, çöpünü konbini ya da otelde at.", "Konbini (7-Eleven)", "Tek Görsel+Not", "Voiceover (TR)", "japonya, gizlituyolar, kultur, traveltips, japonyagezi", 3, "Kolay"),
 ("Gizli Tüyolar", "Onsen'e girmeden bilmen gereken kurallar?", "Ayıp etme 👇", "Onsen'e girmeden önce yıkanılır, havuza sabunsuz-tertemiz girilir. Mayo yok. Küçük havluyu suya sokma, başının üstüne koy. Bazı onsenler dövmeliyi almıyor; önceden kontrol et.", "Stok/Grafik", "Seslendirme", "Voiceover (TR)", "onsen, japonya, gorgukurallari, kultur, traveltips", 2, "Orta"),
 ("Gizli Tüyolar", "Japonya'da sigara nerede içilir?", "Sokakta değil 👇", "Birçok şehir merkezinde sokakta yürürken sigara yasak; sadece belirli 'smoking area'larda içiliyor. İçmek istersen tabelalı alanları ara; kapalı bazı kafelerde ise serbest olabiliyor.", "Stok/Grafik", "Tek Görsel+Not", "Voiceover (TR)", "japonya, gizlituyolar, kultur, traveltips, japonyagezi", 3, "Kolay"),
 # --- Konaklama ---
 ("Konaklama", "Japonya'da otel odaları neden bu kadar küçük?", "Ama pişman olmazsın 👇", "Şehir merkezinde 'business hotel' odaları kompakt tasarlanıyor; küçük ama tertemiz, konforlu ve çok merkezi. Bütçe + konum önemliyse ideal. Geniş oda istiyorsan biraz merkez dışına bak.", "Stok/Grafik", "Seslendirme", "Voiceover (TR)", "japonya, konaklama, otel, japonyagezi, traveltips", 3, "Kolay"),
 ("Konaklama", "Ryokan nedir, denemeli mi?", "En az 1 gece 👇", "Ryokan geleneksel Japon hanı: tatami zemin, futon yatak, çoğunda onsen ve kaiseki (çok kaplı geleneksel yemek). Modern otelden bambaşka bir deneyim; gezinde en az bir gece dene.", "Stok/Grafik", "Montaj", "Sakin lofi", "ryokan, japonya, konaklama, japonyagezi, traveltips", 3, "Orta"),
]

rownum = 4
for i, item in enumerate(K, 1):
    kat, hook, ekran, cevap, gorsel, fmt, ses, tags, oncelik, zorluk = item
    row = [i, kat, hook, ekran, cevap, gorsel, fmt, ses, tags, oncelik, zorluk, "Fikir", "", ""]
    for c in range(1, len(head) + 1):
        cell = ws.cell(row=rownum, column=c, value=row[c - 1])
        cell.alignment = WRAP
        cell.border = BORDER
        if i % 2 == 1:
            cell.fill = PatternFill("solid", fgColor=GRI)
    # öncelik 1 vurgusu
    pc = ws.cell(row=rownum, column=10)
    if oncelik == 1:
        pc.fill = PatternFill("solid", fgColor="C6EFCE")
    pc.alignment = WRAP_CENTER
    ws.cell(row=rownum, column=1).alignment = WRAP_CENTER
    ws.row_dimensions[rownum].height = 66
    rownum += 1

ws.freeze_panes = "A4"
dv_durum = DataValidation(type="list", formula1='"Fikir,Planlandı,Çekildi,Üretildi,Yayınlandı"', allow_blank=True)
ws.add_data_validation(dv_durum)
dv_durum.add(f"L4:L{rownum-1}")
dv_o2 = DataValidation(type="list", formula1='"1,2,3"', allow_blank=True)
ws.add_data_validation(dv_o2)
dv_o2.add(f"J4:J{rownum-1}")

toplam_reel = len(K)
oncelik1 = sum(1 for x in K if x[8] == 1)

# =========================================================
# 04 PAYLAŞIM TAKVİMİ
# =========================================================
ws = wb.create_sheet("04 Paylaşım Takvimi")
ws.sheet_properties.tabColor = "2E86C1"
ws["A1"] = "Paylaşım Takvimi — İlk 8 Hafta (haftada 6)"
ws["A1"].font = TITLE_FONT
ws.merge_cells("A1:F1")
ws["A2"] = "Strateji: İlk 2 hafta elindeki en güçlü çekimlerle 'hızlı başlangıç'. Pzt/Çrş/Cum ana reel, Sal/Prş tüyo, Cmt seri/rota."
ws["A2"].font = SUB_FONT
ws.merge_cells("A2:F2")
head = ["Hafta", "Gün", "Kategori", "Reel (Hook)", "Format", "Not"]
for c, h in enumerate(head, 1):
    ws.cell(row=4, column=c, value=h)
style_header(ws, 4, len(head))
set_widths(ws, [8, 8, 22, 46, 16, 26])

gunler = ["Pzt", "Sal", "Çrş", "Prş", "Cum", "Cmt"]
# Haftalık plan: elde-çekim-var + yüksek etki önce
plan = [
 # Hafta 1 - hepsi öncelik 1, elde çekim var
 ("Nara & Geyikler", "Nara geyikleri neden başını eğiyor?", "Montaj", "Açılış patlaması; save'e oynar"),
 ("Konbini & Yeme-İçme", "Konbini'de bütün geziyi kurtaran 3 şey?", "Montaj", "Pratik, kaydedilir"),
 ("Tokyo Rehberi", "Tokyo'da en iyi manzara bedava — nerede?", "Seslendirme", "Merak + değer"),
 ("Shinkansen Rehberi", "Shinkansen'de Fuji hangi koltuktan görünür?", "Montaj", "Fuji görseli güçlü"),
 ("Tema Parkları", "Tokyo Disney'de sırada beklememenin yolu?", "Montaj", "Geniş kitle"),
 ("Rotalar & Süre", "13 günde Japonya: efsane rota nasıl kurulur?", "Carousel", "Profil sabitleme adayı"),
 # Hafta 2
 ("Osaka Rehberi", "Dotonbori'de o meşhur kareyi nerede çekersin?", "Montaj", "Gece ışık estetiği"),
 ("Kyoto Rehberi", "Fushimi Inari'nin torii'lerinde ne yazıyor?", "Montaj", "Şaşırtan bilgi"),
 ("Alışveriş & Tax-Free", "Japonya'da Uniqlo'yu neden orada vurmalısın?", "Montaj", "Tax-free tüyosu"),
 ("Varışta İlk Yapılacaklar", "Japonya'ya inince ilk 30 dakikada ne yapmalısın?", "Seslendirme", "Yeni gezgin mıknatısı"),
 ("Tema Parkları", "Universal Nintendo World'e nasıl girilir?", "Montaj", "Nintendo görseli"),
 ("Osaka Rehberi", "Osaka Kalesi'nin içinde ne var?", "Montaj", "Kale + havadan"),
 # Hafta 3
 ("Gidiş Rehberi", "Valizini nasıl hafiflettik? (kimse bilmiyor)", "Montaj", "Takkyubin tüyosu"),
 ("Nara & Geyikler", "Nara geyiklerine ne yedirebilirsin?", "Montaj", "Seri #2"),
 ("Kültür & Görgü Kuralları", "Japon tuvaletindeki büyük düğme ne işe yarar?", "Montaj", "Eğlenceli/viral"),
 ("Kyoto Rehberi", "Fushimi Inari'de kalabalıktan nasıl kaçarsın?", "Montaj", "Pratik"),
 ("Tokyo Rehberi", "Shibuya Crossing'i yukarıdan nasıl çekersin?", "Montaj", "İkonik görsel"),
 ("Sezon & Planlama", "Japonya'da 12 ay: hangi ay ne var?", "Carousel", "Kaydet canavarı"),
 # Hafta 4
 ("Gidiş Rehberi", "JR Pass hâlâ mantıklı mı, yoksa para mı yakıyor?", "Seslendirme", "Tartışma yaratır"),
 ("Konbini & Yeme-İçme", "Onigiri paketi nasıl açılır?", "Montaj", "Kısa, tatmin edici"),
 ("Aile ile Japonya", "Çocuklarla Tokyo'da ne yapılır?", "Montaj", "Aile kitlesi"),
 ("Kyoto Rehberi", "Kyoto'da metro neden işe yaramaz?", "Seslendirme", "Yanlış-doğru formatı"),
 ("Tema Parkları", "Universal'de Express Pass şart mı?", "Seslendirme", "Karar rehberi"),
 ("Rotalar & Süre", "Osaka'dan günübirlik gidilecek 6 rota?", "Carousel", "Rota serisi"),
 # Hafta 5-8 : karışık, kalanları ve tekrar-güçlüleri
 ("Alışveriş & Tax-Free", "Japonya'nın en büyük outlet'i hangisi?", "Seslendirme", "Mitsui Outlet"),
 ("Varışta İlk Yapılacaklar", "Japonya'da kart mı geçer nakit mi?", "Montaj", "Pratik"),
 ("Tokyo Rehberi", "Tokyo metrosunda yanlış çıkış seni nasıl yakar?", "Seslendirme", "Zaman kaybı korkusu"),
 ("Uçuş & Bilet", "Çin aktarmalı Japonya uçuşları: nelere dikkat?", "Seslendirme", "Güncel bilgi"),
 ("Osaka Rehberi", "Osaka mı Tokyo mu daha lezzetli?", "Montaj", "Tartışma"),
 ("Gizli Tüyolar", "Japonya'da kaybettiğin cüzdan geri gelir mi?", "Seslendirme", "Şaşırtıcı"),

 ("Kültür & Görgü Kuralları", "Japonya'da yürürken yemek neden ayıp?", "Seslendirme", "Görgü serisi"),
 ("Alışveriş & Tax-Free", "Pokemon Center'da neyi kaçırma?", "Montaj", "Pokemon görseli"),
 ("Shinkansen Rehberi", "NOZOMI, HIKARI, KODAMA farkı ne?", "Seslendirme", "Bilgi"),
 ("Aile ile Japonya", "Çocukla Shinkansen zor mu?", "Montaj", "Aile"),
 ("Sezon & Planlama", "Kiraz çiçeği (sakura) ne zaman açar?", "Tek Görsel+Not", "Sezonluk"),
 ("Rotalar & Süre", "Japonya'ya kaç gün yeter?", "Tek Görsel+Not", "SSS"),

 ("Gidiş Rehberi", "Japonya'ya en ucuz hangi ay gidilir?", "Tek Görsel+Not", "Planlama"),
 ("Konbini & Yeme-İçme", "Japonya'da su almana gerek var mı?", "Tek Görsel+Not", "Tasarruf"),
 ("Tokyo Rehberi", "Tokyo Tower mı Skytree mi?", "Montaj", "Karşılaştırma"),
 ("Ulaşım & Uygulamalar", "Japonya'da hangi uygulamalar şart?", "Seslendirme", "Fayda"),
 ("Gizli Tüyolar", "Onsen'e girmeden bilmen gereken kurallar?", "Seslendirme", "Görgü"),
 ("Kyoto Rehberi", "Gion'da geyşa fotoğrafı çekmek yasak mı?", "Seslendirme", "Uyarı"),

 ("Kültür & Görgü Kuralları", "Japonya'da bahşiş veriliyor mu?", "Tek Görsel+Not", "Kısa bilgi"),
 ("Ulaşım & Uygulamalar", "Valizle gezme: coin locker sırrı", "Montaj", "Pratik"),
 ("Tema Parkları", "Disney'de popcorn kovası neden bu kadar önemli?", "Montaj", "Eğlence"),
 ("Osaka Rehberi", "Osaka'yı gökten görünce ne değişiyor?", "Montaj", "Havadan"),
 ("Gidiş Rehberi", "Japonya'da internet: pocket wifi mi eSIM mi?", "Seslendirme", "Planlama"),
 ("Konaklama", "Ryokan nedir, denemeli mi?", "Montaj", "Deneyim"),
]

r = 5
for wk in range(8):
    for d in range(6):
        idx = wk * 6 + d
        if idx >= len(plan):
            break
        kat, hook, fmt, note = plan[idx]
        vals = [f"Hafta {wk+1}" if d == 0 else "", gunler[d], kat, hook, fmt, note]
        for c in range(1, len(head) + 1):
            cell = ws.cell(row=r, column=c, value=vals[c - 1])
            cell.alignment = WRAP
            cell.border = BORDER
            if wk % 2 == 1:
                cell.fill = PatternFill("solid", fgColor=GRI)
        ws.row_dimensions[r].height = 30
        r += 1
ws.freeze_panes = "A5"

# =========================================================
# 05 ROADMAP
# =========================================================
ws = wb.create_sheet("05 Roadmap")
ws.sheet_properties.tabColor = "6C3483"
ws["A1"] = "Roadmap — Sıfırdan 100K'ya Faz Faz"
ws["A1"].font = TITLE_FONT
ws.merge_cells("A1:F1")
head = ["Faz", "Odak", "İçerik Tipi", "Cadence", "Büyüme Taktiği", "Başarı Ölçütü / Milestone"]
for c, h in enumerate(head, 1):
    ws.cell(row=3, column=c, value=h)
style_header(ws, 3, len(head))
set_widths(ws, [22, 30, 26, 16, 40, 34])
roadmap = [
 ["Faz 0 – Kurulum (1 hafta)", "Kimlik ve altyapı", "Kapak şablonu, bio, öne çıkanlar", "—", "Tutarlı görsel dil (kırmızı bant + @japonyaseyahat), seri isimleri, ilk 9 gridi hazırla.", "Profil yayında, 9 reel hazır kuyrukta"],
 ["Faz 1 – Isınma (Ay 1-2)", "Format bulma", "Kısa montaj + tek-bilgi tüyo", "7/hafta", "Elindeki çekimlerle üret, hook'ları test et, hangi konu tutuyor gözle; en iyi 3 formatı ikiye katla.", "1.000 takipçi, tutan bir 'viral' reel"],
 ["Faz 2 – İvme (Ay 3-5)", "Seri + erişim", "Seriler, trend ses, altyazı", "7/hafta", "Trend sesleri yakala, ekran altyazısı ekle, save/paylaşıma oynayan rehberler; yorumlara reels ile cevap ver.", "10.000 takipçi, düzenli 20K+ izlenme"],
 ["Faz 3 – Otorite (Ay 6-12)", "Marka ve derinlik", "Rehber carousel + iş birliği", "7-10/hafta", "Diğer gezi/hesaplarla iş birliği, tekrar-izlenmeye oynayan kurgu, en iyi serilerin devamı, bio'da rehber/ürün.", "100.000 takipçi, 'Japonya rehberi' otoritesi"],
]
r = write_rows(ws, 4, roadmap, len(head))
ws.freeze_panes = "A4"
for rr in range(4, r):
    ws.row_dimensions[rr].height = 70

# =========================================================
out = Path("content/Japonya_Reels_Roadmap.xlsx")
out.parent.mkdir(exist_ok=True)
wb.save(out)
print(f"Kaydedildi: {out.resolve()}")
print(f"Toplam reel fikri: {toplam_reel} | Öncelik-1 (hemen üretilebilir): {oncelik1}")
print(f"Sayfalar: {wb.sheetnames}")
